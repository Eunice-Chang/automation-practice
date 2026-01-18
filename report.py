import sys  # sys：拿命令行参数 sys.argv
import csv  # csv：读 CSV（把文本表格变成字典行）
import os   # os：检查文件是否存在
import statistics
import argparse
from openpyxl import Workbook # openpyxl.Workbook：创建 Excel（.xlsx）
from openpyxl.styles import Font, PatternFill, Alignment  # styles 相关：设置字体、填充色、对齐、边框
from openpyxl.utils import get_column_letter  # get_column_letter：把列号 1/2/3 转成 A/B/C（用于设置列宽）
from openpyxl.chart import BarChart, Reference # BarChart：柱状图

# ===样式函数===
def auto_fit_columns(ws, min_width=10, max_width=40): #自动设置列宽
    """Auto-fit column widths based on cell text lengths."""
    for col_idx in range(1, ws.max_column + 1): # 当前 sheet 一共有多少列
        max_len = 0
        col_letter = get_column_letter(col_idx) # col_idx = 1 → col_letter = "A"
        for row_idx in range(1, ws.max_row + 1): # 当前列有多少行
            v = ws.cell(row=row_idx, column=col_idx).value #取出单元格的值
            if v is None:  
                continue #跳过空单元格
            s = str(v) #把内容统一变成字符串
            if len(s) > max_len:
                max_len = len(s)
        # rough width estimate
        width = max(min_width, min(max_width, max_len + 2))
        ws.column_dimensions[col_letter].width = width # excel的列宽设置 ws.column_dimensions["A"].width = 20

def style_header(ws, header_row=1): #设置表头格式 style_header-变量名，header_row=1-第一行为表头
    """Bold + center header row, with a light fill."""
    header_font = Font(bold=True) # 字体样式-加粗
    header_fill = PatternFill("solid", fgColor="EDEDED") #背景颜色 solid-纯色填充，EDEDED-浅灰色
    header_align = Alignment(horizontal="center", vertical="center") #对齐方式

    for cell in ws[header_row]: #遍历表头行中的每个单元格
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align

def style_details_pass_fail(ws, status_col=3, start_row=2): # 根据Status染色整行
    """Color entire row based on Status."""
    fill_pass = PatternFill("solid", fgColor="E8F5E9")  # light green
    fill_fail = PatternFill("solid", fgColor="FFEBEE")  # light red

    for r in range(start_row, ws.max_row + 1):
        status = ws.cell(row=r, column=status_col).value
        if status == "Pass":
            fill = fill_pass
        elif status == "Fail":
            fill = fill_fail
        else:
            continue

        for c in range(1, ws.max_column + 1):
            ws.cell(row=r, column=c).fill = fill

def set_freeze_panes(ws, cell="A2"): #冻结窗格，让第一行表头永远可见 冻结A2单元格“上方 + 左侧”的区域
    ws.freeze_panes = cell

def apply_rate_color(cell, rate, good=80, warn=60):
    """Apply background color to a cell based on rate value."""
    fill_good = PatternFill("solid", fgColor="E8F5E9")   # green
    fill_warn = PatternFill("solid", fgColor="FFF8E1")   # yellow
    fill_bad  = PatternFill("solid", fgColor="FFEBEE")   # red

    if rate >= good:
        cell.fill = fill_good
    elif rate >= warn:
        cell.fill = fill_warn
    else:
        cell.fill = fill_bad

def apply_median_color(cell, median_score, pass_score):
    fill_median = PatternFill("solid", fgColor="FFCDD2")

    if median_score < pass_score:
        cell.fill = fill_median


def parse_args():
    parser = argparse.ArgumentParser(
        description="Generate Excel report from score CSV"
    )

    parser.add_argument("input_csv", help="Input CSV file")
    parser.add_argument("output_xlsx", help="Output Excel file")

    parser.add_argument(
        "--pass-score",
        type=float,
        default=60,
        help="Pass score threshold (default: 60)"
    )

    parser.add_argument(
        "--chart",
        action="store_true",
        help="Generate Pass/Fail chart"
    )

    return parser.parse_args()

def create_pass_fail_chart(ws, row_pass, row_fail):
    chart = BarChart()
    chart.title = "Pass / Fail Summary"
    chart.y_axis.title = "Count"
    chart.x_axis.title = "Result"

    data = Reference(
        ws,
        min_col=2,
        min_row=row_pass,
        max_row=row_fail
    )

    categories = Reference(
        ws,
        min_col=1,
        min_row=row_pass,
        max_row=row_fail
    )

    chart.add_data(data, titles_from_data=False)
    chart.set_categories(categories)

    ws.add_chart(chart, "D2")



def run(): #工具入口

    args = parse_args()

    input_file = args.input_csv
    output_file = args.output_xlsx
    pass_score = args.pass_score

    if not os.path.exists(input_file):
        raise FileNotFoundError(input_file)

    # ===== 2. 读取 CSV（并校验列名）=====
    with open(input_file, "r", newline="", encoding="utf-8") as f:
        reader = csv.DictReader(f)

        required_cols = {"name", "score"}
        if not required_cols.issubset(reader.fieldnames):
            raise ValueError("CSV must contain columns: name, score")

        rows = list(reader)

    # ===== 3. 处理数据 =====
    details = []
    scores = []
    pass_count = 0

    for row in rows:
        name = row.get("name", "")

        # 分数允许整数/小数（更通用）
        try:
            score = float(row.get("score", ""))
        except ValueError:
            raise ValueError(
                f"Score '{row.get('score')}' for '{name}' is not a number."
            )

        scores.append(score)

        status = "Pass" if score >= pass_score else "Fail"
        if status == "Pass":
            pass_count += 1

        details.append([name, score, status])

    total = len(scores)

    if total == 0:
        raise ValueError("No valid data rows found in CSV file.")

    average = sum(scores) / total
    median_score = statistics.median(scores)
    max_score = max(scores)
    min_score = min(scores)
    pass_rate = pass_count / total * 100

    # ===== 4. 创建 Excel =====
    wb = Workbook()

    # ---- Sheet 1：明细 ----
    ws_details = wb.active
    ws_details.title = "Details"

    ws_details.append(["Name", "Score", "Status"])
    for r in details:
        ws_details.append(r)

    # 基础样式
    style_header(ws_details, header_row=1)
    set_freeze_panes(ws_details, "A2")

    # 数字格式（分数显示两位小数，你也可以改成 0 位）
    for r in range(2, ws_details.max_row + 1):
        ws_details.cell(row=r, column=2).number_format = "0.00"

    # Pass/Fail 行上色
    style_details_pass_fail(ws_details, status_col=3, start_row=2)

    # 自动列宽
    auto_fit_columns(ws_details)

    # ---- Sheet 2：统计 ----
    ws_summary = wb.create_sheet(title="Summary")
    ws_summary.append(["Metric", "Value"])
    ws_summary.append(["Pass Score", pass_score])
    ws_summary.append(["Total", total])
    ws_summary.append(["Average", average])
    ws_summary.append(["Median", median_score])
    row_median = ws_summary.max_row
    ws_summary.append(["Max", max_score])
    ws_summary.append(["Min", min_score])
    ws_summary.append(["Pass", pass_count])
    row_pass = ws_summary.max_row
    ws_summary.append(["Fail", total - pass_count])
    row_fail = ws_summary.max_row
    ws_summary.append(["Pass Count", pass_count])
    ws_summary.append(["Pass Rate (%)", pass_rate])
    row_pass_rate = ws_summary.max_row

    style_header(ws_summary, header_row=1)
    set_freeze_panes(ws_summary, "A2")
    auto_fit_columns(ws_summary)

    # Summary 数字格式更好看
    # Pass Score / Average / Median / Max / Min：两位小数
    for row_idx in [2, 4, row_median , 6, 7]:
        ws_summary.cell(row=row_idx, column=2).number_format = "0.00"
    # Pass Rate：两位小数
    ws_summary.cell(row=row_pass_rate, column=2).number_format = "0.00"

    # 对齐：Summary 左列左对齐，右列右对齐
    for r in range(2, ws_summary.max_row + 1):
        ws_summary.cell(row=r, column=1).alignment = Alignment(horizontal="left")
        ws_summary.cell(row=r, column=2).alignment = Alignment(horizontal="right")

    apply_rate_color(ws_summary.cell(row=row_pass_rate, column=2), pass_rate)

    apply_median_color(ws_summary.cell(row=row_median,column=2), median_score, pass_score)
    
    if args.chart:
        create_pass_fail_chart(ws_summary, row_pass, row_fail)

    # ===== 5. 保存 =====
    wb.save(output_file)

    print("Styled Excel report generated successfully.")
    print("Output:", output_file)

def main(): # 排除 输入 CSV 不存在 / 路径错误
    try:
        run()
    except FileNotFoundError as e:
        print(f"Error: file not found -> {e.filename}")
    except PermissionError:
        print("Error: output file is open. Please close the Excel file and retry.")
    except ValueError as e:
        print(f"Error: {e}")
    

if __name__ == "__main__":
    main()
